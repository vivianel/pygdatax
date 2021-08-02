import nxlib
import xeuss
import openpyxl as xl
from os.path import exists
import argparse


def run(excelFile):
    wb = xl.load_workbook(excelFile)
    sheet = wb[wb.sheetnames[0]]
    mask_file = sheet.cell(7, 2).value
    dark_file = sheet.cell(8, 2).value
    ec_file = sheet.cell(9, 2).value
    eb_file = sheet.cell(10, 2).value
    if dark_file is not None:
        nxlib.build_nexus_from_edf(dark_file)
        dark_file = dark_file.split('.')[0] + '.nxs'
    if ec_file is not None:
        nxlib.build_nexus_from_edf(ec_file)
        ec_file = ec_file.split('.')[0] + '.nxs'
    if eb_file is not None:
        nxlib.build_nexus_from_edf(eb_file)
        eb_file = eb_file.split('.')[0] + '.nxs'

    nrows = sheet.max_row
    files = []
    thicks = []
    for i in range(12, nrows+1):
        file = sheet.cell(i, 1).value
        thick = sheet.cell(i, 2).value
        if exists(file):
            files.append(file)
            thicks.append(thick)
        else:
            sheet.cell(i, 2).value = 'Not found'
    wb.save(excelFile)
    for file in files:
        nxlib.build_nexus_from_edf(file)
    for i in range(len(files)):
        files[i] = files[i].split('.')[0]+'.nxs'
    x0 = sheet.cell(2, 2).value
    if x0 is not None:
        x0 = float(sheet.cell(2, 2).value)
    y0 = sheet.cell(3, 2).value
    if y0 is not None:
        y0 = float(sheet.cell(3, 2).value)
    distance = sheet.cell(4, 2).value
    if distance is not None:
        distance = float(sheet.cell(4, 2).value)
    nbins = sheet.cell(5, 2).value
    if nbins is not None:
        nbins = int(sheet.cell(5, 2).value)
    if dark_file is not None:
        xeuss.set_beam_center(dark_file, x0=x0, y0=y0, new_entry=False)  # direct_beam_file=directbeam, new_entry=False)
        xeuss.azimutal_integration(dark_file, bins=nbins, mask=mask_file)
    if ec_file is not None:
        xeuss.set_beam_center(ec_file, x0=x0, y0=y0, new_entry=False)  # direct_beam_file=directbeam, new_entry=False)
        xeuss.azimutal_integration(ec_file, bins=nbins, mask=mask_file)
    if eb_file is not None:
        xeuss.set_beam_center(eb_file, x0=x0, y0=y0, new_entry=False)  # direct_beam_file=directbeam, new_entry=False)
        xeuss.azimutal_integration(eb_file, bins=nbins, mask=mask_file)

    for file, ep in zip(files, thicks):
        # try:
        xeuss.set_beam_center(file, x0=x0, y0=y0, new_entry=False)  # direct_beam_file=directbeam, new_entry=False)
        xeuss.azimutal_integration(file, bins=nbins, mask=mask_file)
        xeuss.resu(file, dark_file=dark_file, ec_file=ec_file, eb_file=eb_file,
                   distance=distance, thickness=ep)
        # except (TypeError, ValueError, ZeroDivisionError, NameError):
        #     pass


def get_nexus_files(excelFile):
    wb = xl.load_workbook(excelFile)
    sheet = wb[wb.sheetnames[0]]
    nrows = sheet.max_row
    files = []
    for i in range(12, nrows + 1):
        file = sheet.cell(i, 1).value
        nxfile = file.replace('.edf', '.nxs')
        if exists(nxfile):
            files.append(nxfile)
    return files


def main():
    parser = argparse.ArgumentParser(description='azimutal averaging,substraction and normalization of SAXS data : pygdataredauto \'file path\'')
    parser.add_argument("excel_file", help='excel file for data treatment that must be within the data folder',
                        type=str)
    args = parser.parse_args()
    run(args.excel_file)


if __name__ == '__main__':
    # file = '/home/achennev/Documents/xeuss/fc aout 2020/pyred_auto.xlsx'

    main()




